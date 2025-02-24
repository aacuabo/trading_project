import pandas as pd
import sqlite3
from openpyxl import load_workbook
from datetime import datetime, timedelta
from dateutil import parser
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    filename='streamlit_app.log',
    filemode='a'
)
logger = logging.getLogger(__name__)

def add_column_if_not_exists(conn, table_name, column_name):
    """Add a new column to the specified table if it doesn't exist"""
    cursor = conn.cursor()
    cursor.execute(f"PRAGMA table_info({table_name})")
    columns = [info[1] for info in cursor.fetchall()]
    if column_name not in columns:
        conn.execute(f'ALTER TABLE {table_name} ADD COLUMN "{column_name}" REAL')
        conn.commit()

def generate_intervals(base_date):
    """Generate 5-minute intervals for a 24-hour period"""
    intervals = []
    start_time = base_date.replace(hour=0, minute=5, second=0)
    for i in range(288):  # 5-minute intervals for 24 hours
        interval_time = start_time + timedelta(minutes=5 * i)
        intervals.append(interval_time.strftime('%Y-%m-%d %H:%M:%S'))
    return intervals

def process_sheet_data(sheet, base_date, conn, sheet_name):
    """Process individual sheet data and return DataFrame for preview"""
    intervals = generate_intervals(base_date)

    # Add new column if it doesn't exist (moved outside loop for efficiency)
    add_column_if_not_exists(conn, 'MQ', sheet_name)

    # Check existing intervals (moved outside loop for efficiency)
    cursor = conn.cursor()
    placeholders = ','.join(['?' for _ in intervals])
    cursor.execute(f'''
        SELECT "Interval", "{sheet_name}"
        FROM MQ
        WHERE "Interval" IN ({placeholders})
    ''', intervals)
    existing_data = {row[0]: row[1] for row in cursor.fetchall()}

    # Lists to hold data for DataFrame creation
    interval_list = []
    date_list = []
    time_list = []
    hour_list = []
    value_list = []

    data_index = 0
    batch_updates = []
    batch_inserts = []

    for row_num in range(13, 37):
        for col_num in range(5, 50, 4):
            if data_index < 288:
                interval = intervals[data_index]
                value = sheet.cell(row=row_num, column=col_num).value
                value = float(value) if value not in (None, '') else 0.0

                timestamp = datetime.strptime(interval, '%Y-%m-%d %H:%M:%S')
                date_str = timestamp.strftime('%Y-%m-%d')
                time_str = timestamp.strftime('%H:%M:%S')
                hour_str = (
                    (timestamp - timedelta(minutes=5)).strftime('%Y-%m-%d %H:00:00')
                    if timestamp.strftime('%H:%M:%S') != '00:00:00'
                    else (timestamp - timedelta(hours=1)).strftime('%Y-%m-%d %H:00:00')
                )

                interval_list.append(interval)
                date_list.append(date_str)
                time_list.append(time_str)
                hour_list.append(hour_str)
                value_list.append(value)

                data_index += 1

    processed_df = pd.DataFrame({
        'Interval': interval_list,
        'Date': date_list,
        'Time': time_list,
        'Hour': hour_list,
        sheet_name: value_list  # Sheet name as column
    })

    for index, row in processed_df.iterrows():
        interval_val = row['Interval']
        value_val = row[sheet_name]

        if interval_val in existing_data:
            batch_updates.append((value_val, interval_val))
        else:
            batch_inserts.append(tuple(row[['Interval', 'Date', 'Time', 'Hour', sheet_name]].tolist())) # Ensure correct order

    return batch_updates, batch_inserts, processed_df # Return preview_df


def update_total_mq(conn):
    """Update Total_MQ column for all rows"""
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(MQ)")
    columns = [info[1] for info in cursor.fetchall()]
    numeric_cols = [col for col in columns if col not in ['Interval', 'Date', 'Time', 'Hour', 'Total_MQ']]

    update_total_sql = f'''
        UPDATE MQ
        SET "Total_MQ" = ({' + '.join([f'COALESCE("{col}", 0)' for col in numeric_cols])})
    '''
    conn.execute(update_total_sql)
    conn.commit()
    return numeric_cols

def update_hourly_data(conn, numeric_cols):
    """Update MQ_Hourly table with aggregated data"""
    cursor = conn.cursor()

    # Get hourly aggregated data
    cursor.execute("""
        SELECT Hour as Interval,
                MAX(Date) as Date,
                substr(Hour, -8) as Time,  -- Extract time portion (HH:MM:SS) from Hour
                """ +
                ','.join([f'SUM(COALESCE("{col}", 0)) as "{col}"' for col in numeric_cols]) +
                f',SUM(COALESCE("Total_MQ", 0)) as "Total_MQ"' +
        """ FROM MQ
        GROUP BY Hour
    """)

    hourly_data = cursor.fetchall()
    column_names = [description[0] for description in cursor.description]

    # Ensure columns exist in MQ_Hourly
    for col in numeric_cols + ['Total_MQ']:
        add_column_if_not_exists(conn, 'MQ_Hourly', col)

    # Process hourly data with UPSERT logic
    for row in hourly_data:
        row_dict = dict(zip(column_names, row))
        interval = row_dict['Interval']

        insert_cols = ['Interval', 'Date', 'Time'] + numeric_cols + ['Total_MQ']
        columns = ', '.join([f'"{col}"' for col in insert_cols])
        placeholders = ', '.join(['?'] * len(insert_cols))

        insert_values = (interval, row_dict['Date'], row_dict['Time']) + \
                        tuple(row_dict[col] for col in numeric_cols + ['Total_MQ'])

        conn.execute(f'''
            INSERT OR REPLACE INTO MQ_Hourly ({columns})
            VALUES ({placeholders})
        ''', insert_values)

    conn.commit()

def process_mq_reports(file_paths, db_path):
    """Main function to process MQ reports and return preview DataFrame"""
    preview_dfs = [] # List to hold preview DataFrames from each sheet
    success = False  # Flag to track overall processing success
    try:
        conn = sqlite3.connect(db_path)

        # Process each file
        for file_path in file_paths:
            wb = load_workbook(file_path, data_only=True)

            # Process each sheet in the workbook
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                base_date = parser.parse(sheet['B2'].value)

                # Process sheet data and get preview_df
                batch_updates, batch_inserts, preview_df = process_sheet_data(sheet, base_date, conn, sheet_name)
                preview_dfs.append(preview_df) # Add preview_df to the list

                # Execute batch updates
                if batch_updates:
                    conn.executemany(f'''
                        UPDATE MQ
                        SET "{sheet_name}" = ?
                        WHERE "Interval" = ?
                    ''', batch_updates)

                # Execute batch inserts
                if batch_inserts:
                    conn.executemany(f'''
                        INSERT INTO MQ ("Interval", "Date", "Time", "Hour", "{sheet_name}")
                        VALUES (?, ?, ?, ?, ?)
                    ''', batch_inserts)

                conn.commit()
            success = True # Set success to True if at least one file is processed without error

        # Update Total_MQ and hourly data
        numeric_cols = update_total_mq(conn)
        update_hourly_data(conn, numeric_cols)

        conn.close()
        return pd.concat(preview_dfs, ignore_index=True) if preview_dfs else pd.DataFrame(), success # Return concatenated preview DataFrame and success status

    except Exception as e:
        if 'conn' in locals():
            conn.close()
        logger.error(f"Error processing MQ reports: {str(e)}")
        return pd.DataFrame(), False # Return empty DataFrame and failure status in case of error